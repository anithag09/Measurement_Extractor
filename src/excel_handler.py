import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ExcelHandler:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.workbook = None
        
    def read_questions(self):
        """Read questions from Excel file and organize by page"""
        # Read Excel without headers, using column indices
        df = pd.read_excel(self.excel_path, header=None)
        
        questions = {}
        current_page = None
        
        # Iterate through rows, using column index 1 (B column)
        for _, row in df.iterrows():
            # Check if the cell is not empty
            if pd.notna(row[1]):  # Column B has index 1
                text = str(row[1]).strip()
                
                if text.startswith('Page-'):
                    current_page = int(text.split('-')[1])
                    questions[current_page] = []
                elif current_page is not None and text:
                    questions[current_page].append(text)
        
        return questions
    
    def update_answers(self, page_num, results):
        """Update answers in Excel file"""
        if self.workbook is None:
            self.workbook = load_workbook(self.excel_path)
        ws = self.workbook.active
        
        # Define colors for success and error
        success_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        error_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        # Find the starting row for the page
        start_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=2).value  # Column B
            if cell_value and str(cell_value).strip() == f'Page-{page_num}':
                start_row = row + 1  # Start from next row after page header
                break
        
        if start_row is None:
            raise ValueError(f"Page-{page_num} not found in Excel file")
        
        # Write results
        for i, result in enumerate(results):
            cell = ws.cell(row=start_row + i, column=3)  # Column C
            
            # Handle different types of values
            if isinstance(result.get('value'), (list, dict)):
                cell.value = str(result['value'])  # Convert complex types to string
            else:
                cell.value = result.get('value', "Not found")
            
            cell.fill = success_fill if result.get('value') is not None else error_fill
        
        # Save after each update
        self.workbook.save(self.excel_path)
    
    def close(self):
        """Close the workbook and save changes"""
        if self.workbook is not None:
            self.workbook.save(self.excel_path)
            self.workbook.close()